"""
Enhanced Grading Export System - Comprehensive data export for professor assessment
"""

import sqlite3
import pandas as pd
import streamlit as st
import json
from datetime import datetime, timedelta
from typing import Dict, List, Any, Optional
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from components.database import DATABASE_PATH

class GradingExportSystem:
    """Enhanced system for exporting student interactions and assessments for grading"""
    
    def __init__(self):
        self.database_path = DATABASE_PATH
        
        # Grade scale mapping
        self.grade_mapping = {
            (90, 100): "A",
            (85, 89): "A-", 
            (80, 84): "B+",
            (77, 79): "B",
            (73, 76): "B-",
            (70, 72): "C+",
            (67, 69): "C",
            (63, 66): "C-",
            (60, 62): "D+",
            (57, 59): "D",
            (0, 56): "F"
        }
    
    def export_comprehensive_grading_report(self, article_title: str = None, 
                                          week_number: int = None) -> Dict[str, Any]:
        """Create comprehensive grading report with multiple sheets/sections"""
        
        # Get student sessions
        sessions_df = self._get_student_sessions(article_title, week_number)
        
        if sessions_df.empty:
            return {"error": "No student sessions found for the specified criteria."}
        
        # Generate detailed analysis
        student_summaries = self._generate_student_summaries(sessions_df)
        interaction_details = self._get_detailed_interactions(sessions_df)
        quality_analysis = self._analyze_interaction_quality(sessions_df)
        discussion_insights = self._extract_discussion_insights(sessions_df)
        
        return {
            "student_summaries": student_summaries,
            "interaction_details": interaction_details,
            "quality_analysis": quality_analysis,
            "discussion_insights": discussion_insights,
            "export_timestamp": datetime.now().isoformat(),
            "criteria": {
                "article_title": article_title,
                "week_number": week_number
            }
        }
    
    def export_to_excel(self, grading_data: Dict[str, Any], filename: str = None) -> bytes:
        """Export grading data to Excel format with multiple sheets"""
        
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"grading_report_{timestamp}.xlsx"
        
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Sheet 1: Student Summary
        self._create_student_summary_sheet(wb, grading_data["student_summaries"])
        
        # Sheet 2: Quality Analysis
        self._create_quality_analysis_sheet(wb, grading_data["quality_analysis"])
        
        # Sheet 3: Detailed Interactions
        self._create_interactions_sheet(wb, grading_data["interaction_details"])
        
        # Sheet 4: Discussion Insights
        self._create_discussion_insights_sheet(wb, grading_data["discussion_insights"])
        
        # Sheet 5: Grading Rubric
        self._create_rubric_sheet(wb)
        
        # Save to bytes
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer.getvalue()
    
    def export_individual_student_report(self, student_id: str, 
                                       article_title: str = None) -> Dict[str, Any]:
        """Generate detailed report for individual student"""
        
        conn = sqlite3.connect(self.database_path)
        
        # Get student sessions
        query = """
            SELECT s.session_id, s.article_title, s.start_time, s.duration_minutes,
                   s.message_count, s.max_level_reached,
                   GROUP_CONCAT(m.content || ' [' || m.role || ']', ' | ') as conversation
            FROM chat_sessions s
            LEFT JOIN chat_messages m ON s.session_id = m.session_id
            WHERE s.user_id = ? AND s.user_type = 'Student'
        """
        
        params = [student_id]
        if article_title:
            query += " AND s.article_title = ?"
            params.append(article_title)
        
        query += " GROUP BY s.session_id ORDER BY s.start_time DESC"
        
        try:
            sessions_df = pd.read_sql_query(query, conn, params=params)
            
            if sessions_df.empty:
                return {"error": f"No sessions found for student {student_id}"}
            
            # Analyze individual performance
            performance_analysis = self._analyze_individual_performance(sessions_df, student_id)
            
            # Get conversation transcripts
            transcripts = self._format_conversation_transcripts(sessions_df)
            
            # Generate recommendations
            recommendations = self._generate_student_recommendations(performance_analysis)
            
            return {
                "student_id": student_id,
                "performance_analysis": performance_analysis,
                "transcripts": transcripts,
                "recommendations": recommendations,
                "generated_at": datetime.now().isoformat()
            }
            
        except Exception as e:
            return {"error": f"Error generating report for {student_id}: {str(e)}"}
        finally:
            conn.close()
    
    def _get_student_sessions(self, article_title: str = None, 
                            week_number: int = None) -> pd.DataFrame:
        """Get student sessions with filtering options"""
        
        conn = sqlite3.connect(self.database_path)
        
        query = """
            SELECT s.session_id, s.user_id, s.article_title, s.start_time,
                   s.duration_minutes, s.message_count, s.max_level_reached,
                   a.week_number, a.learning_objectives
            FROM chat_sessions s
            LEFT JOIN articles a ON s.article_title = a.title
            WHERE s.user_type = 'Student'
        """
        
        params = []
        if article_title:
            query += " AND s.article_title = ?"
            params.append(article_title)
        
        if week_number:
            query += " AND a.week_number = ?"
            params.append(week_number)
        
        query += " ORDER BY s.start_time DESC"
        
        try:
            return pd.read_sql_query(query, conn, params=params)
        except Exception as e:
            st.error(f"Error retrieving sessions: {e}")
            return pd.DataFrame()
        finally:
            conn.close()
    
    def _generate_student_summaries(self, sessions_df: pd.DataFrame) -> List[Dict[str, Any]]:
        """Generate summary statistics for each student"""
        
        summaries = []
        
        for student_id in sessions_df['user_id'].unique():
            student_sessions = sessions_df[sessions_df['user_id'] == student_id]
            
            # Calculate metrics
            total_sessions = len(student_sessions)
            total_duration = student_sessions['duration_minutes'].sum()
            avg_duration = student_sessions['duration_minutes'].mean()
            total_messages = student_sessions['message_count'].sum()
            avg_messages = student_sessions['message_count'].mean()
            max_level = student_sessions['max_level_reached'].max()
            avg_level = student_sessions['max_level_reached'].mean()
            
            # Calculate engagement score (0-100)
            engagement_score = min(100, (
                (avg_duration / 30 * 25) +  # Duration component (30 min = full points)
                (avg_messages / 20 * 25) +  # Message count component
                (max_level / 4 * 25) +      # Cognitive level component
                (total_sessions / 3 * 25)   # Session frequency component
            ))
            
            # Assign letter grade
            letter_grade = self._score_to_letter_grade(engagement_score)
            
            summaries.append({
                "student_id": student_id,
                "total_sessions": int(total_sessions),
                "total_duration_minutes": round(total_duration, 1),
                "avg_duration_minutes": round(avg_duration, 1),
                "total_messages": int(total_messages),
                "avg_messages_per_session": round(avg_messages, 1),
                "max_cognitive_level": int(max_level),
                "avg_cognitive_level": round(avg_level, 1),
                "engagement_score": round(engagement_score, 1),
                "letter_grade": letter_grade,
                "articles_discussed": list(student_sessions['article_title'].unique())
            })
        
        # Sort by engagement score (highest first)
        summaries.sort(key=lambda x: x['engagement_score'], reverse=True)
        
        return summaries
    
    def _get_detailed_interactions(self, sessions_df: pd.DataFrame) -> List[Dict[str, Any]]:
        """Get detailed interaction data for each session"""
        
        conn = sqlite3.connect(self.database_path)
        interactions = []
        
        for _, session in sessions_df.iterrows():
            # Get messages for this session
            query = """
                SELECT role, content, timestamp, message_order
                FROM chat_messages
                WHERE session_id = ?
                ORDER BY message_order
            """
            
            try:
                messages_df = pd.read_sql_query(query, conn, params=[session['session_id']])
                
                if not messages_df.empty:
                    # Separate student and AI messages
                    student_messages = messages_df[messages_df['role'] == 'student']
                    ai_messages = messages_df[messages_df['role'] == 'assistant']
                    
                    # Analyze message quality
                    quality_metrics = self._calculate_session_quality_metrics(student_messages)
                    
                    interactions.append({
                        "session_id": session['session_id'],
                        "student_id": session['user_id'],
                        "article_title": session['article_title'],
                        "start_time": session['start_time'],
                        "duration_minutes": session['duration_minutes'],
                        "student_message_count": len(student_messages),
                        "ai_message_count": len(ai_messages),
                        "max_level_reached": session['max_level_reached'],
                        "quality_metrics": quality_metrics,
                        "conversation_sample": self._get_conversation_sample(messages_df)
                    })
            
            except Exception as e:
                continue
        
        conn.close()
        return interactions
    
    def _analyze_interaction_quality(self, sessions_df: pd.DataFrame) -> Dict[str, Any]:
        """Analyze overall quality patterns across all interactions"""
        
        conn = sqlite3.connect(self.database_path)
        
        # Get all student messages
        session_ids = sessions_df['session_id'].tolist()
        if not session_ids:
            return {}
        
        placeholders = ','.join(['?'] * len(session_ids))
        query = f"""
            SELECT m.session_id, m.content, s.user_id, s.article_title
            FROM chat_messages m
            JOIN chat_sessions s ON m.session_id = s.session_id
            WHERE m.role = 'user' AND m.session_id IN ({placeholders})
        """
        
        try:
            messages_df = pd.read_sql_query(query, conn, params=session_ids)
            
            # Analyze quality indicators
            quality_analysis = {
                "total_student_messages": len(messages_df),
                "avg_message_length": messages_df['content'].str.len().mean(),
                "question_frequency": self._calculate_question_frequency(messages_df),
                "critical_thinking_indicators": self._identify_critical_thinking(messages_df),
                "concept_mentions": self._count_concept_mentions(messages_df),
                "spatial_reasoning_frequency": self._analyze_spatial_reasoning(messages_df)
            }
            
            return quality_analysis
            
        except Exception as e:
            return {"error": f"Error analyzing quality: {str(e)}"}
        finally:
            conn.close()
    
    def _extract_discussion_insights(self, sessions_df: pd.DataFrame) -> Dict[str, Any]:
        """Extract insights for class discussion preparation"""
        
        conn = sqlite3.connect(self.database_path)
        
        session_ids = sessions_df['session_id'].tolist()
        if not session_ids:
            return {}
        
        placeholders = ','.join(['?'] * len(session_ids))
        query = f"""
            SELECT m.content, s.user_id, s.article_title
            FROM chat_messages m
            JOIN chat_sessions s ON m.session_id = s.session_id
            WHERE m.role = 'user' AND m.session_id IN ({placeholders})
        """
        
        try:
            messages_df = pd.read_sql_query(query, conn, params=session_ids)
            
            # Extract discussion-worthy content
            insights = {
                "interesting_questions": self._extract_interesting_questions(messages_df),
                "common_misconceptions": self._identify_misconceptions(messages_df),
                "emerging_themes": self._identify_themes(messages_df),
                "student_insights": self._extract_student_insights(messages_df),
                "concept_connections": self._identify_concept_connections(messages_df)
            }
            
            return insights
            
        except Exception as e:
            return {"error": f"Error extracting insights: {str(e)}"}
        finally:
            conn.close()
    
    def _score_to_letter_grade(self, score: float) -> str:
        """Convert numerical score to letter grade"""
        for (min_score, max_score), grade in self.grade_mapping.items():
            if min_score <= score <= max_score:
                return grade
        return "F"
    
    def _calculate_session_quality_metrics(self, student_messages: pd.DataFrame) -> Dict[str, float]:
        """Calculate quality metrics for a single session"""
        
        if student_messages.empty:
            return {}
        
        total_messages = len(student_messages)
        all_text = ' '.join(student_messages['content'].astype(str))
        
        # Calculate metrics
        avg_message_length = student_messages['content'].str.len().mean()
        questions_asked = sum('?' in msg for msg in student_messages['content'])
        question_rate = questions_asked / total_messages if total_messages > 0 else 0
        
        # Word count and complexity
        total_words = sum(len(msg.split()) for msg in student_messages['content'])
        avg_words_per_message = total_words / total_messages if total_messages > 0 else 0
        
        return {
            "avg_message_length": round(avg_message_length, 1),
            "question_rate": round(question_rate, 2),
            "avg_words_per_message": round(avg_words_per_message, 1),
            "total_questions": questions_asked,
            "message_count": total_messages
        }
    
    def _get_conversation_sample(self, messages_df: pd.DataFrame) -> str:
        """Get a sample of the conversation for review"""
        
        if messages_df.empty:
            return ""
        
        # Get first 3 exchanges (up to 6 messages)
        sample_messages = messages_df.head(6)
        
        conversation = []
        for _, msg in sample_messages.iterrows():
            role = "Student" if msg['role'] == 'student' else "AI"
            content = msg['content'][:100] + "..." if len(msg['content']) > 100 else msg['content']
            conversation.append(f"{role}: {content}")
        
        return "\n".join(conversation)
    
    def _calculate_question_frequency(self, messages_df: pd.DataFrame) -> float:
        """Calculate frequency of questions in student messages"""
        
        if messages_df.empty:
            return 0.0
        
        question_count = sum('?' in msg for msg in messages_df['content'])
        return question_count / len(messages_df)
    
    def _identify_critical_thinking(self, messages_df: pd.DataFrame) -> Dict[str, int]:
        """Identify critical thinking indicators in messages"""
        
        critical_thinking_keywords = [
            "however", "although", "because", "therefore", "suggests", 
            "implies", "evidence", "assume", "evaluate", "critique"
        ]
        
        indicators = {}
        all_text = ' '.join(messages_df['content'].astype(str)).lower()
        
        for keyword in critical_thinking_keywords:
            indicators[keyword] = all_text.count(keyword)
        
        return indicators
    
    def _count_concept_mentions(self, messages_df: pd.DataFrame) -> Dict[str, int]:
        """Count mentions of key landscape ecology concepts"""
        
        concepts = [
            "fragmentation", "connectivity", "scale", "edge effect", 
            "patch", "corridor", "habitat", "landscape", "gis", "spatial"
        ]
        
        concept_counts = {}
        all_text = ' '.join(messages_df['content'].astype(str)).lower()
        
        for concept in concepts:
            concept_counts[concept] = all_text.count(concept)
        
        return concept_counts
    
    def _analyze_spatial_reasoning(self, messages_df: pd.DataFrame) -> Dict[str, float]:
        """Analyze spatial reasoning indicators"""
        
        spatial_keywords = ["spatial", "scale", "resolution", "pattern", "distribution", 
                          "connectivity", "buffer", "distance", "proximity", "network"]
        
        all_text = ' '.join(messages_df['content'].astype(str)).lower()
        total_messages = len(messages_df)
        
        spatial_mentions = sum(all_text.count(keyword) for keyword in spatial_keywords)
        spatial_frequency = spatial_mentions / total_messages if total_messages > 0 else 0
        
        return {
            "spatial_mentions_total": spatial_mentions,
            "spatial_frequency_per_message": round(spatial_frequency, 2)
        }
    
    def _extract_interesting_questions(self, messages_df: pd.DataFrame) -> List[str]:
        """Extract questions that might be good for class discussion"""
        
        questions = []
        for msg in messages_df['content']:
            if '?' in msg and len(msg.split()) > 8:  # Substantive questions
                questions.append(msg[:200] + "..." if len(msg) > 200 else msg)
        
        return questions[:10]  # Top 10 questions
    
    def _identify_misconceptions(self, messages_df: pd.DataFrame) -> List[str]:
        """Identify potential misconceptions for class discussion"""
        
        # This is a simplified version - would use more sophisticated NLP in practice
        misconception_patterns = [
            "all fragmentation is bad", "bigger is always better", 
            "edges are always harmful", "gis shows the truth"
        ]
        
        misconceptions = []
        all_text = ' '.join(messages_df['content'].astype(str)).lower()
        
        for pattern in misconception_patterns:
            if pattern in all_text:
                misconceptions.append(f"Some students may think: {pattern}")
        
        return misconceptions
    
    def _identify_themes(self, messages_df: pd.DataFrame) -> Dict[str, int]:
        """Identify emerging themes in student discussions"""
        
        themes = {
            "habitat_fragmentation": ["fragment", "patch", "isolation"],
            "scale_effects": ["scale", "resolution", "hierarchy"],
            "connectivity": ["connect", "corridor", "movement", "dispersal"],
            "human_impacts": ["human", "urban", "agriculture", "development"],
            "conservation": ["protect", "conserve", "manage", "restore"]
        }
        
        theme_counts = {}
        all_text = ' '.join(messages_df['content'].astype(str)).lower()
        
        for theme, keywords in themes.items():
            count = sum(all_text.count(keyword) for keyword in keywords)
            theme_counts[theme] = count
        
        return theme_counts
    
    def _extract_student_insights(self, messages_df: pd.DataFrame) -> List[Dict[str, str]]:
        """Extract notable student insights"""
        
        insights = []
        insight_indicators = ["i think", "i believe", "it seems", "this suggests", "because"]
        
        for _, row in messages_df.iterrows():
            msg = row['content'].lower()
            if any(indicator in msg for indicator in insight_indicators) and len(msg.split()) > 10:
                insights.append({
                    "student_id": row['user_id'],
                    "insight": row['content'][:300] + "..." if len(row['content']) > 300 else row['content']
                })
        
        return insights[:15]  # Top 15 insights
    
    def _identify_concept_connections(self, messages_df: pd.DataFrame) -> List[str]:
        """Identify when students make connections between concepts"""
        
        connection_indicators = ["relates to", "connects to", "similar to", "like", "also"]
        connections = []
        
        for msg in messages_df['content']:
            msg_lower = msg.lower()
            if any(indicator in msg_lower for indicator in connection_indicators) and len(msg.split()) > 8:
                connections.append(msg[:200] + "..." if len(msg) > 200 else msg)
        
        return connections[:10]
    
    def _create_student_summary_sheet(self, workbook: Workbook, summaries: List[Dict[str, Any]]):
        """Create Excel sheet with student summaries"""
        
        ws = workbook.create_sheet("Student Summary")
        
        # Headers
        headers = ["Student ID", "Sessions", "Total Duration (min)", "Avg Duration (min)", 
                  "Total Messages", "Avg Messages", "Max Level", "Avg Level", 
                  "Engagement Score", "Grade", "Articles"]
        
        ws.append(headers)
        
        # Format headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Data rows
        for summary in summaries:
            ws.append([
                summary["student_id"],
                summary["total_sessions"],
                summary["total_duration_minutes"],
                summary["avg_duration_minutes"],
                summary["total_messages"],
                summary["avg_messages_per_session"],
                summary["max_cognitive_level"],
                summary["avg_cognitive_level"],
                summary["engagement_score"],
                summary["letter_grade"],
                ", ".join(summary["articles_discussed"])
            ])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_quality_analysis_sheet(self, workbook: Workbook, quality_data: Dict[str, Any]):
        """Create Excel sheet with quality analysis"""
        
        ws = workbook.create_sheet("Quality Analysis")
        
        # Overall metrics
        ws.append(["Quality Metrics", "Value"])
        ws.append(["Total Student Messages", quality_data.get("total_student_messages", 0)])
        ws.append(["Average Message Length", quality_data.get("avg_message_length", 0)])
        ws.append(["Question Frequency", quality_data.get("question_frequency", 0)])
        
        ws.append([])  # Empty row
        
        # Critical thinking indicators
        ws.append(["Critical Thinking Indicators"])
        ct_indicators = quality_data.get("critical_thinking_indicators", {})
        for indicator, count in ct_indicators.items():
            ws.append([indicator, count])
        
        ws.append([])  # Empty row
        
        # Concept mentions
        ws.append(["Concept Mentions"])
        concept_mentions = quality_data.get("concept_mentions", {})
        for concept, count in concept_mentions.items():
            ws.append([concept, count])
    
    def _create_interactions_sheet(self, workbook: Workbook, interactions: List[Dict[str, Any]]):
        """Create Excel sheet with detailed interactions"""
        
        ws = workbook.create_sheet("Detailed Interactions")
        
        # Headers
        headers = ["Session ID", "Student ID", "Article", "Date", "Duration (min)", 
                  "Student Messages", "Max Level", "Sample Conversation"]
        
        ws.append(headers)
        
        # Data rows
        for interaction in interactions:
            ws.append([
                interaction["session_id"],
                interaction["student_id"],
                interaction["article_title"],
                interaction["start_time"],
                interaction["duration_minutes"],
                interaction["student_message_count"],
                interaction["max_level_reached"],
                interaction["conversation_sample"]
            ])
    
    def _create_discussion_insights_sheet(self, workbook: Workbook, insights: Dict[str, Any]):
        """Create Excel sheet with discussion insights"""
        
        ws = workbook.create_sheet("Discussion Insights")
        
        # Interesting questions
        ws.append(["Interesting Student Questions"])
        questions = insights.get("interesting_questions", [])
        for question in questions:
            ws.append([question])
        
        ws.append([])  # Empty row
        
        # Student insights
        ws.append(["Notable Student Insights"])
        student_insights = insights.get("student_insights", [])
        for insight in student_insights:
            ws.append([f"{insight.get('student_id', 'Unknown')}: {insight.get('insight', '')}"])
    
    def _create_rubric_sheet(self, workbook: Workbook):
        """Create Excel sheet with grading rubric"""
        
        ws = workbook.create_sheet("Grading Rubric")
        
        ws.append(["Landscape Ecology Chatbot Interaction Rubric"])
        ws.append([])
        
        # Criteria
        criteria = [
            ["Depth of Analysis (20%)", "Critical thinking, original insights"],
            ["Concept Integration (15%)", "Connects multiple concepts"],
            ["Question Quality (15%)", "Probing, thoughtful questions"],
            ["Engagement Consistency (15%)", "Sustained participation"],
            ["Cognitive Progression (15%)", "Reaches higher thinking levels"],
            ["Spatial Reasoning (20%)", "GIS and spatial thinking skills"]
        ]
        
        ws.append(["Criteria", "Description"])
        for criterion in criteria:
            ws.append(criterion)
        
        ws.append([])
        ws.append(["Grade Scale:"])
        ws.append(["A: 90-100", "Excellent"])
        ws.append(["B: 80-89", "Good"]) 
        ws.append(["C: 70-79", "Satisfactory"])
        ws.append(["D: 60-69", "Needs Improvement"])
        ws.append(["F: Below 60", "Unsatisfactory"])
    
    def _analyze_individual_performance(self, sessions_df: pd.DataFrame, student_id: str) -> Dict[str, Any]:
        """Analyze individual student performance"""
        
        performance = {
            "total_sessions": len(sessions_df),
            "total_duration": sessions_df['duration_minutes'].sum(),
            "avg_duration": sessions_df['duration_minutes'].mean(),
            "progression": self._track_cognitive_progression(sessions_df),
            "engagement_pattern": self._analyze_engagement_pattern(sessions_df),
            "strengths": [],
            "areas_for_improvement": []
        }
        
        # Identify strengths and areas for improvement
        if performance["avg_duration"] > 20:
            performance["strengths"].append("Sustained engagement in conversations")
        else:
            performance["areas_for_improvement"].append("Increase depth of engagement")
        
        max_level = sessions_df['max_level_reached'].max()
        if max_level >= 3:
            performance["strengths"].append("Reaches advanced cognitive levels")
        else:
            performance["areas_for_improvement"].append("Work toward synthesis and evaluation levels")
        
        return performance
    
    def _track_cognitive_progression(self, sessions_df: pd.DataFrame) -> List[int]:
        """Track cognitive level progression over time"""
        
        sessions_sorted = sessions_df.sort_values('start_time')
        return sessions_sorted['max_level_reached'].tolist()
    
    def _analyze_engagement_pattern(self, sessions_df: pd.DataFrame) -> Dict[str, Any]:
        """Analyze student's engagement patterns"""
        
        sessions_sorted = sessions_df.sort_values('start_time')
        
        return {
            "session_frequency": len(sessions_df),
            "duration_trend": "improving" if sessions_sorted['duration_minutes'].iloc[-1] > sessions_sorted['duration_minutes'].iloc[0] else "declining",
            "consistency": sessions_df['duration_minutes'].std()
        }
    
    def _format_conversation_transcripts(self, sessions_df: pd.DataFrame) -> List[Dict[str, Any]]:
        """Format conversation transcripts for review"""
        
        transcripts = []
        
        for _, session in sessions_df.iterrows():
            if pd.notna(session.get('conversation')):
                conversation_parts = session['conversation'].split(' | ')
                formatted_conversation = []
                
                for i, part in enumerate(conversation_parts):
                    if ' [user]' in part:
                        formatted_conversation.append(f"Student: {part.replace(' [user]', '')}")
                    elif ' [assistant]' in part:
                        formatted_conversation.append(f"AI: {part.replace(' [assistant]', '')}")
                
                transcripts.append({
                    "session_id": session['session_id'],
                    "article_title": session['article_title'],
                    "start_time": session['start_time'],
                    "transcript": "\n".join(formatted_conversation)
                })
        
        return transcripts
    
    def _generate_student_recommendations(self, performance: Dict[str, Any]) -> List[str]:
        """Generate recommendations for student improvement"""
        
        recommendations = []
        
        if performance["total_sessions"] < 3:
            recommendations.append("Increase frequency of chatbot interactions to deepen understanding")
        
        if performance["avg_duration"] < 15:
            recommendations.append("Spend more time exploring concepts through extended dialogue")
        
        progression = performance.get("progression", [])
        if progression and max(progression) < 3:
            recommendations.append("Challenge yourself to reach synthesis and evaluation cognitive levels")
        
        if not recommendations:
            recommendations.append("Continue excellent engagement! Consider exploring more complex connections between concepts")
        
        return recommendations